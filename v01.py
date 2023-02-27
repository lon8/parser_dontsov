import datetime
import json
import requests
import openpyxl
import xlsxwriter
from bs4 import BeautifulSoup
from operator import itemgetter
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options

proxies = {
    'http': 'http://20.111.54.16:80',
    'http': 'http://152.67.73.106:80',
    'http': 'http://23.227.39.154:80',
    'http': 'http://23.227.39.93:80',
    'http': 'http://141.193.213.96:80',
    'http': 'http://103.21.244.195:80',
    'http': 'http://172.64.149.43:80',
    'http': 'http://172.64.159.235:80',
    'http': 'http://172.67.180.20:80',
    'http': 'http://203.28.8.223:80'
}


def write_headers():
    worksheet.write(0, 0, 'Brand')
    worksheet.write(0, 1, 'Article')
    worksheet.write(0, 2, "Provider")
    worksheet.write(0, 3, "Date of issue")
    worksheet.write(0, 4, "Availability")
    worksheet.write(0, 5, "Price")

location = 'Прайс опт полный прайс1.xlsx'

book = openpyxl.load_workbook(location)
sheet = book.active

workbook = xlsxwriter.Workbook(f'parse_{datetime.datetime.today()}.xlsx')
worksheet = workbook.add_worksheet()
write_headers()

counter = 1
for ar in range(11, 21): #1803
    is_one = False
    article = sheet.cell(row=ar, column=2).value
    print("ARTICLE ----->", article)

    worksheet.write(counter, 0, sheet.cell(row=ar, column=1).value)
    worksheet.write(counter, 1, article)

    responce = requests.get(f'https://v01.ru/auto/search/?q={article}', proxies=proxies)

    soup = BeautifulSoup(responce.text, 'lxml')
    
    header = soup.find('tr', class_='htitle no-hover').text
    table = soup.find('table')
    if header == "Искомый артикул в наличии":
        is_one = True
        rows = table.find_all('tr', class_='hproduct no-hover catalog-product-row todayDeliver supplierIsStock hlast')
    else:
        rows = table.find_all('tr', class_='hproduct no-hover hlast')
    
    # Проходимся по строчкам
    count = 1
    supplier_objects = {}
    for row in rows:
        # Берем дату
        date_first = row.find('div', class_='icon_info fas fa-home toolttip-fade') \
            .get('title') \
            .replace('\n                            ', '') \
            .replace('\n                        ', '')
        datesoup = BeautifulSoup(date_first, 'lxml')
        date_str = datesoup.find('b').text
        date = datetime.datetime.strptime(date_str, "%H:%M %d.%m.%Y")
        # Берем поставщика
        provider = row.find('td', class_='supplier').text.replace('\n', '')
        # Берем наличие
        availability = row.find('td', class_='instock text-center').text.replace('\n', '').replace(' ', '')
        # Если в наличии меньше двух, то пропускаем
        availability_int = int(''.join([x for x in availability if x.isdigit()]))
        # Берем цену
        price = int(row.find('td', class_='price number-cell')['title'])
        # Записываем элемент в словарь
        supplier_objects[count] = {"Date": date, 'Supplier': provider, "Availability":availability_int, "Price": price}
        count += 1
    # Сортируем по дате и цене
    supplier_objects = dict(sorted(supplier_objects.items(), key=lambda x: (x[1]['Date'], x[1]['Price'], -x[1]["Availability"])))
    keys = list(supplier_objects.keys())
    final_supplier = supplier_objects[1]
    final_date = final_supplier['Date'].strftime("%H:%M %d.%m.%Y")

    worksheet.write(counter, 2, final_supplier['Supplier'])
    worksheet.write(counter, 3, final_date)
    worksheet.write(counter, 4, final_supplier['Availability'])
    worksheet.write(counter, 5, final_supplier['Price'])

    counter += 1

workbook.close()