from bs4 import BeautifulSoup
import requests
import threading
import openpyxl
import queue
from fp.fp import FreeProxy
import random
import xlsxwriter
import datetime
import time

proxy = []
for i in range(1, 10):
    try:
        prox = FreeProxy(https=True, rand=True).get()
        proxy.append(prox)
        print('Добавили прокси')
    except: break




def write_headers():
    worksheet.write(0, 0, 'Brand')
    worksheet.write(0, 1, 'Article')
    worksheet.write(0, 2, "Provider")
    worksheet.write(0, 3, "Date of issue")
    worksheet.write(0, 4, "Availability")
    worksheet.write(0, 5, "Price")

location = 'pop.xlsx'


# Первая строчка XLSX - файла
MIN_ROW = 11

# Последняя строчка XLSX - файла
MAX_ROW = 1803

book = openpyxl.load_workbook(location)
sheet = book.active

workbook = xlsxwriter.Workbook(f'parse_{datetime.datetime.today()}.xlsx')
worksheet = workbook.add_worksheet()
write_headers()

articles = []

for iter in range(MIN_ROW, MAX_ROW + 1):
    artic = sheet.cell(row=iter, column=2).value
    brand = sheet.cell(row=iter, column=1).value
    articles.append(artic)
# Закрываем файл
book.close()

counter = 1

def kernel(table, header):
    global counter
    if header == "Искомый артикул в наличии":
        is_one = True
        rows = table.find_all('tr', class_='hproduct no-hover catalog-product-row todayDeliver supplierIsStock hlast')
    else:
        rows =  table.find_all('tr')
        rows.pop(0)
        rows.pop(0)
    # Проходимся по строчкам
    count = 1
    supplier_objects = {}
    first_row = True
    brand = ''
    for row in rows:
        date_first = row.find('div', class_='icon_info fas fa-home toolttip-fade') \
            .get('title') \
            .replace('\n                            ', '') \
            .replace('\n                        ', '')
        datesoup = BeautifulSoup(date_first, 'lxml')
        date_str = datesoup.find('b').text
        date = datetime.datetime.strptime(date_str, "%H:%M %d.%m.%Y")
        # Берем артикул
        if(first_row is True):
            article = row.find('td', class_='sku catalog-table-rowspan').find('a').text.replace(' ', '').replace('\n', '')
            # Берем бренд
            brand = row.find('td', class_='brand catalog-table-rowspan')["title"]
        # Берем поставщика
        provider = row.find('td', class_='supplier').text.replace('\n', '')
        # Берем наличие
        availability = row.find('td', class_='instock text-center').text.replace('\n', '').replace(' ', '')
        # Если в наличии меньше двух, то пропускаем
        availability_int = int(''.join([x for x in availability if x.isdigit()]))
        # Берем цену
        price = int(row.find('td', class_='price number-cell')['title'])
        # Записываем элемент в словарь
        supplier_objects[count] = {"Date": date, 'Supplier': provider, "Availability":availability_int, "Price": price}
        first_row = False
        count += 1
    # Сортируем по дате и цене
    supplier_objects = dict(sorted(supplier_objects.items(), key=lambda x: (x[1]['Date'], x[1]['Price'], -x[1]["Availability"])))
    keys = list(supplier_objects.keys())
    final_supplier = supplier_objects[keys[0]]
    final_date = final_supplier['Date'].strftime("%H:%M %d.%m.%Y")
    worksheet.write(counter, 0, brand)
    worksheet.write(counter, 1, article)
    worksheet.write(counter, 2, final_supplier['Supplier'])
    worksheet.write(counter, 3, final_date)
    worksheet.write(counter, 4, final_supplier['Availability'])
    worksheet.write(counter, 5, final_supplier['Price'])


    counter += 1




def make_request(art, proxy):
    global counter

    for art in articles:
        try:
            # proxies = {
            #     "https": random.choice(proxy)
            # }
            time.sleep(1)
            responce = requests.get(f'https://v01.ru/auto/search/?q={art}', proxies={'https': proxy})
            print("\n\n", "----->", f'https://v01.ru/auto/search/?q={art}', '\n\n')
            soup = BeautifulSoup(responce.text, 'lxml')
            header = soup.find('tr', class_='htitle no-hover').text
            table = soup.find('table')
            kernel(table, header)
        except:
            products = soup.find_all('tr', class_='catalog-table-brand-row js-search-item')
            for product in products:
                href = product['onclick'].split("'")[1::2]
                link = "https://v01.ru" + href[0]
                print("\n\n", "----->", link, '\n\n')
                responce_req = requests.get(link, proxies={'https': proxy})
                time.sleep(1)
                soup = BeautifulSoup(responce_req.text, 'lxml')
                try:
                    header = soup.find('tr', class_='htitle no-hover').text
                except:
                    print('Товар не найден в каталоге')
                    continue
                table = soup.find('table')
                kernel(table, header)
    

def worker(q, proxy):
    while True:
        art = q.get()
        make_request(art, proxy)
        q.task_done()

concurrent_requests = len(proxy)
q = queue.Queue()
for article in articles:
    q.put(article)

threads = []
for i in range(concurrent_requests):
    thread = threading.Thread(target=worker, args=(q, proxy[i-1]))
    threads.append(thread)
    thread.start()

for thread in threads:
    thread.join()

workbook.close()